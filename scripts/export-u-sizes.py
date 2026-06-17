"""Export U size lookup by outcome and rating delta from calculator W column logic."""
import json
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "temp-gameplay.xlsx"
OUT = ROOT / "u-size-lookup.json"

OUTCOMES = ["HR", "3B", "2B", "1B", "IF1B", "BB", "FO", "PO", "K"]


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    calc = wb["calculator"]

    lookup = {}
    row_by_code = {
        "HR": 11,
        "3B": 12,
        "2B": 13,
        "1B": 14,
        "IF1B": 15,
        "BB": 16,
        "FO": 17,
        "PO": 18,
        "K": 19,
    }

    for code, row in row_by_code.items():
        size = calc.cell(row, 23).value
        delta = calc.cell(row, 22).value
        lookup[code] = {
            "defaultDelta": delta,
            "defaultSize": size,
        }

    lookup["GO"] = {"defaultSize": calc.cell(20, 23).value}
    lookup["LO"] = {"defaultSize": calc.cell(21, 23).value}

    OUT.write_text(json.dumps(lookup, indent=2), encoding="utf-8")
    print(json.dumps(lookup, indent=2))


if __name__ == "__main__":
    main()
